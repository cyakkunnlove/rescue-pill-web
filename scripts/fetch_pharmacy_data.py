#!/usr/bin/env python3
"""Build the pharmacy search JSON from the current MHLW workbook."""

from __future__ import annotations

from datetime import datetime, timezone
from hashlib import sha256
from html.parser import HTMLParser
from io import BytesIO
import json
import os
from pathlib import Path
import re
import ssl
import sys
import unicodedata
from urllib.parse import urljoin, urlparse
from urllib.request import Request, urlopen

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "openpyxl is required. Run: python3 -m pip install -r requirements-data.txt",
        file=sys.stderr,
    )
    raise SystemExit(1)


PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_FILE = PROJECT_ROOT / "public/data/otc_pharmacies.json"
METADATA_FILE = PROJECT_ROOT / "public/data/otc_pharmacies.meta.json"

SOURCE_PAGE = "https://www.mhlw.go.jp/stf/kinnkyuuhininnyaku_00005.html"
SOURCE_FILE_OVERRIDE = os.environ.get("MHLW_PHARMACY_XLSX_URL")
SOURCE_UPDATED_AT_OVERRIDE = os.environ.get("MHLW_PHARMACY_UPDATED_AT")

EXPECTED_HEADERS = (
    "薬局等番号",
    "都道府県",
    "薬局等名称",
    "住所",
    "電話番号",
    "女性",
    "男性",
    "答えたくない",
    "HP",
    "開局時間",
    "時間外対応の有無",
    "時間外の電話番号",
    "プライバシー確保策",
    "事前連絡",
    "備考",
    "削除日",
)


class SourcePageParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[dict[str, object]] = []
        self.text_parts: list[str] = []
        self.current_link: dict[str, object] | None = None

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() != "a":
            return
        href = dict(attrs).get("href")
        if href:
            self.current_link = {"href": href, "text": []}

    def handle_data(self, data: str) -> None:
        self.text_parts.append(data)
        if self.current_link is not None:
            link_text = self.current_link["text"]
            if isinstance(link_text, list):
                link_text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "a" and self.current_link is not None:
            self.links.append(self.current_link)
            self.current_link = None


def validate_source_url(source_file: str) -> None:
    parsed = urlparse(source_file)
    if parsed.scheme != "https" or parsed.hostname != "www.mhlw.go.jp":
        raise RuntimeError("Source must be an HTTPS URL on www.mhlw.go.jp")
    if not parsed.path.lower().endswith(".xlsx"):
        raise RuntimeError("Source must be an XLSX workbook")


def download_source_page() -> str:
    request = Request(SOURCE_PAGE, headers={"User-Agent": "RescuePillDataUpdater/1.0"})
    with urlopen(request, timeout=60, context=ssl.create_default_context()) as response:
        if response.status != 200:
            raise RuntimeError(f"MHLW source page download failed: HTTP {response.status}")
        body = response.read()
        charset = response.headers.get_content_charset() or "utf-8"
    if len(body) < 1_000:
        raise RuntimeError("MHLW source page is unexpectedly small")
    return body.decode(charset, errors="strict")


def parse_source_updated_at(page_text: str) -> str:
    normalized = unicodedata.normalize("NFKC", page_text)
    era_dates = re.findall(
        r"令和\s*(元|\d+)\s*年\s*(\d+)\s*月\s*(\d+)\s*日\s*更新",
        normalized,
    )
    parsed_dates: set[str] = set()
    for era_year, month, day in era_dates:
        year = 2019 if era_year == "元" else 2018 + int(era_year)
        parsed_dates.add(f"{year:04d}-{int(month):02d}-{int(day):02d}")

    if not parsed_dates:
        western_dates = re.findall(
            r"(20\d{2})\s*年\s*(\d+)\s*月\s*(\d+)\s*日\s*更新",
            normalized,
        )
        parsed_dates.update(
            f"{int(year):04d}-{int(month):02d}-{int(day):02d}"
            for year, month, day in western_dates
        )

    if len(parsed_dates) != 1:
        raise RuntimeError(
            "Could not identify exactly one update date on the MHLW source page"
        )
    return parsed_dates.pop()


def discover_current_source(page_html: str) -> tuple[str, str]:
    parser = SourcePageParser()
    parser.feed(page_html)
    candidates: list[str] = []
    for link in parser.links:
        href = link["href"]
        link_text = link["text"]
        if not isinstance(href, str) or not isinstance(link_text, list):
            continue
        text = normalize("".join(str(part) for part in link_text))
        source_file = urljoin(SOURCE_PAGE, href)
        parsed = urlparse(source_file)
        if (
            parsed.path.lower().endswith(".xlsx")
            and "要指導医薬品" in text
            and "緊急避妊薬" in text
            and "薬局" in text
            and "一覧" in text
            and "薬剤師確認用" not in text
        ):
            candidates.append(source_file)

    unique_candidates = sorted(set(candidates))
    if len(unique_candidates) != 1:
        raise RuntimeError(
            "Could not identify exactly one current pharmacy XLSX on the MHLW page"
        )

    source_file = unique_candidates[0]
    validate_source_url(source_file)
    source_updated_at = parse_source_updated_at(" ".join(parser.text_parts))
    return source_file, source_updated_at


def resolve_source() -> tuple[str, str]:
    if bool(SOURCE_FILE_OVERRIDE) != bool(SOURCE_UPDATED_AT_OVERRIDE):
        raise RuntimeError(
            "MHLW_PHARMACY_XLSX_URL and MHLW_PHARMACY_UPDATED_AT must be set together"
        )
    if SOURCE_FILE_OVERRIDE and SOURCE_UPDATED_AT_OVERRIDE:
        validate_source_url(SOURCE_FILE_OVERRIDE)
        if not re.fullmatch(r"20\d{2}-\d{2}-\d{2}", SOURCE_UPDATED_AT_OVERRIDE):
            raise RuntimeError("MHLW_PHARMACY_UPDATED_AT must use YYYY-MM-DD")
        return SOURCE_FILE_OVERRIDE, SOURCE_UPDATED_AT_OVERRIDE

    return discover_current_source(download_source_page())


def as_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKC", as_text(value))
    text = re.sub(r"\s+", "", text)
    return re.sub(r"[‐‑‒–—―−]", "-", text)


def location_key(name: object, address: object) -> str:
    return f"{normalize(name)}\0{normalize(address)}"


def to_count(value: object) -> int:
    try:
        return int(float(as_text(value)))
    except ValueError:
        return 0


def existing_coordinates() -> dict[str, dict[str, float]]:
    if not OUTPUT_FILE.exists():
        return {}
    current = json.loads(OUTPUT_FILE.read_text(encoding="utf-8"))
    return {
        location_key(item.get("n"), item.get("a")): {
            "lat": item["lat"],
            "lon": item["lon"],
        }
        for item in current
        if isinstance(item.get("lat"), (int, float))
        and isinstance(item.get("lon"), (int, float))
    }


def download_workbook(source_file: str) -> bytes:
    validate_source_url(source_file)

    request = Request(source_file, headers={"User-Agent": "RescuePillDataUpdater/1.0"})
    with urlopen(request, timeout=60, context=ssl.create_default_context()) as response:
        if response.status != 200:
            raise RuntimeError(f"MHLW download failed: HTTP {response.status}")
        body = response.read()
    if len(body) < 10_000:
        raise RuntimeError("Downloaded workbook is unexpectedly small")
    return body


def find_data_sheet(workbook):
    for worksheet in workbook.worksheets:
        for row_number, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=20, values_only=True), start=1
        ):
            normalized_row = tuple(normalize(cell) for cell in row)
            normalized_headers = tuple(normalize(header) for header in EXPECTED_HEADERS)
            for column_offset in range(max(1, len(row) - len(EXPECTED_HEADERS) + 1)):
                if (
                    normalized_row[
                        column_offset : column_offset + len(EXPECTED_HEADERS)
                    ]
                    == normalized_headers
                ):
                    return worksheet, row_number, column_offset
    raise RuntimeError("Could not find the expected pharmacy-list header")


def parse_pharmacies(body: bytes):
    workbook = load_workbook(BytesIO(body), read_only=True, data_only=True)
    worksheet, header_row, column_offset = find_data_sheet(workbook)
    coordinates = existing_coordinates()
    pharmacies: list[dict[str, object]] = []
    deleted_count = 0

    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        values = list(row[column_offset : column_offset + 16])
        values += [None] * max(0, 16 - len(values))
        if not any(as_text(cell) for cell in values):
            continue
        if as_text(values[15]):
            deleted_count += 1
            continue

        facility_id = as_text(values[0])
        prefecture = as_text(values[1])
        name = as_text(values[2])
        address = as_text(values[3])
        if not facility_id or not prefecture or not name:
            continue

        coordinate = coordinates.get(location_key(name, address), {})
        pharmacies.append(
            {
                "i": facility_id,
                "n": name,
                "p": prefecture,
                "a": address,
                "t": as_text(values[4]),
                "h": as_text(values[9]),
                "e": 1 if as_text(values[10]) == "有" else 0,
                "x": as_text(values[11]),
                "f": to_count(values[5]),
                "m": to_count(values[6]),
                "w": as_text(values[8]),
                "r": 1 if as_text(values[13]) == "要" else 0,
                "lat": coordinate.get("lat"),
                "lon": coordinate.get("lon"),
            }
        )

    if len(pharmacies) < 10_000:
        raise RuntimeError(
            f"Refusing to replace data: only {len(pharmacies)} active facilities parsed"
        )
    return pharmacies, deleted_count, worksheet.title


def atomic_write(path: Path, content: str) -> None:
    temporary = path.with_suffix(f"{path.suffix}.tmp")
    temporary.write_text(content, encoding="utf-8")
    temporary.replace(path)


def main() -> None:
    source_file, source_updated_at = resolve_source()
    print(f"Official page reports update: {source_updated_at}")
    print(f"Downloading: {source_file}")
    body = download_workbook(source_file)
    pharmacies, deleted_count, sheet_name = parse_pharmacies(body)
    rendered_data = json.dumps(
        pharmacies, ensure_ascii=False, separators=(",", ":")
    )
    coordinate_count = sum(
        1
        for item in pharmacies
        if isinstance(item["lat"], (int, float))
        and isinstance(item["lon"], (int, float))
    )
    metadata = {
        "sourcePage": SOURCE_PAGE,
        "sourceFile": source_file,
        "sourceUpdatedAt": source_updated_at,
        "importedAt": datetime.now(timezone.utc).isoformat(),
        "sheetName": sheet_name,
        "total": len(pharmacies),
        "deletedRowsExcluded": deleted_count,
        "coordinateCount": coordinate_count,
        "sourceSha256": sha256(body).hexdigest(),
        "dataSha256": sha256(rendered_data.encode("utf-8")).hexdigest(),
    }

    # Publish metadata first and data second. The client verifies the data hash,
    # so an interrupted two-file update fails closed instead of showing a mixed set.
    atomic_write(
        METADATA_FILE,
        json.dumps(metadata, ensure_ascii=False, indent=2) + "\n",
    )
    atomic_write(OUTPUT_FILE, rendered_data)
    print(
        f"Updated {len(pharmacies):,} pharmacies "
        f"({coordinate_count:,} with coordinates; "
        f"{deleted_count:,} deleted rows excluded)."
    )
    print(f"Source SHA-256: {metadata['sourceSha256']}")
    print(f"Data SHA-256: {metadata['dataSha256']}")


if __name__ == "__main__":
    main()
